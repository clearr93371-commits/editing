# app.py  –  Chef's on Fire | Budget Generator
# Paste this file into a GitHub repo together with requirements.txt,
# then connect the repo at share.streamlit.io to deploy.

import io
import re
from collections import Counter, OrderedDict, defaultdict

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import math
import pandas as pd
import numpy as np

from openpyxl.chart import (BarChart, PieChart, Reference, Series)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation

# ── page config (must be first) ───────────────────────────────────────────────
st.set_page_config(
    page_title="Chef's on Fire – Budget Generator",
    page_icon="🔥",
    layout="wide",
)

# ── styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Archivo+Black&family=Inter:wght@400;500;600&display=swap');

html, body, [class*="css"]          { font-family: 'Inter', sans-serif; }

/* sidebar */
[data-testid="stSidebar"]           { background: #111820; }
[data-testid="stSidebar"] *         { color: #d4c9b8 !important; }
[data-testid="stSidebar"] label     { font-size:.75rem; letter-spacing:.07em;
                                      text-transform:uppercase; color:#7a6e62 !important; }

/* main canvas */
.main .block-container              { background:#f5f1ea; max-width:860px;
                                      padding:2.5rem 2rem 4rem; }

/* wordmark */
.wordmark                           { font-family:'Archivo Black',sans-serif;
                                      font-size:1.75rem; color:#1a1008;
                                      letter-spacing:-.02em; margin-bottom:.15rem; }
.tagline                            { color:#7a6e62; font-size:.85rem;
                                      margin-bottom:2.2rem; }

/* fire button */
.stButton > button                  { background:#bf3b0a; color:#fff; border:none;
                                      border-radius:3px; font-weight:600;
                                      letter-spacing:.04em; padding:.55rem 1.5rem; }
.stButton > button:hover            { background:#9e300a; }

/* section label */
h3                                  { font-family:'Archivo Black',sans-serif;
                                      font-size:1.05rem; color:#1a1008;
                                      margin-top:2rem; letter-spacing:-.01em; }

/* notice strip */
.notice { background:#ede7db; border-left:3px solid #bf3b0a;
          padding:.7rem 1rem; border-radius:0 3px 3px 0;
          font-size:.83rem; color:#3d2f1e; margin-bottom:1.5rem; }

/* metrics */
[data-testid="metric-container"]    { background:#ede7db; border-radius:4px;
                                      padding:.6rem .9rem; }
</style>
""", unsafe_allow_html=True)

# ── header ────────────────────────────────────────────────────────────────────
st.markdown('<p class="wordmark">🔥 Chef\'s on Fire</p>', unsafe_allow_html=True)
st.markdown('<p class="tagline">Budget Generator — internal planning tool</p>',
            unsafe_allow_html=True)

# ── constants & parser ───────────────────────────────────────────────
COLS = ['situacao','pagamento','descricao','custo_unit','unit','dias',
        'custo_total','iva','total_com_iva','patrocinio','entidade',
        'NA ADJUDICAÇÃO','ATÉ O EVENTO',
        'd30','d60','obs','valor_total']

CAT_FILL = 'FFBAD2D4'
EXCLUDED_HEADER_TEXT = ('TOTAL', 'MARGEM DE ERRO', 'INCOME', 'CAPEX', 'SALDO FINAL')

def is_cat_row(desc_cell, desc):
    return desc_cell.font.bold and desc.upper() not in EXCLUDED_HEADER_TEXT

def is_maincat_row(desc_cell):
    fill = desc_cell.fill.fgColor.rgb if desc_cell.fill and desc_cell.fill.fgColor else None
    return 'category' if fill == CAT_FILL else 'subcategory'

def to_num(v, default):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        v = v.strip().replace(',', '.')
        try:
            return float(v)
        except ValueError:
            return default
    return default

def parse_sheet(ws, sheet_name):
    items = []
    current_category    = None
    current_subcategory = None
    
    header_found = False
    data_start_row = 1
    for r_check in range(1, ws.max_row + 1):
        # Look for 'DESCRIÇÃO' in column C (index 2 for 0-based, or 3 for 1-based)
        desc_header_cell_val = ws.cell(r_check, 3).value
        if desc_header_cell_val and 'DESCRIÇÃO' in str(desc_header_cell_val).upper():
            data_start_row = r_check + 1 # Data starts *after* the header row
            header_found = True
            break

    if not header_found:
        data_start_row = 1 
        
    for r in range(data_start_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(COLS) + 1)]
        row  = dict(zip(COLS, vals))
        desc = row['descricao']
        
        if desc is None or str(desc).strip() == '':
            continue
        desc = str(desc).strip()
        desc_cell = ws.cell(r, 3)
        
            
        if is_cat_row(desc_cell, desc):
            if is_maincat_row(desc_cell) == 'category':
                current_category = desc
                current_subcategory = None
            elif is_maincat_row(desc_cell) == 'subcategory':
                current_subcategory = desc
            continue

        # if desc.upper() in EXCLUDED_HEADER_TEXT:
        #     continue
            
        unit_cost = row['custo_unit']
        if unit_cost is None:
            continue
        try:
            unit_cost = float(unit_cost)
        except (TypeError, ValueError):
            continue
        qty       = to_num(row['unit'], 1)
        days      = to_num(row['dias'],  1)
        total_est = to_num(row['custo_total'], unit_cost * qty * days)
        items.append({
            'sheet':             sheet_name,
            'category':          current_category,
            'subcategory':       current_subcategory,
            'description':       desc,
            'unit_cost':         unit_cost,
            'qty':               qty,
            'days':              days,
            'cost_total_pretax': total_est,
            'payment_status':    row['situacao'],
        })
    return items


# ── CELL 11 – ratio validation ────────────────────────────────────────────────
def norm(desc):
    d = desc.lower()
    return d.split(' - ')[0].split(' (')[0].strip()

def build_validated_ratios(all_items):
    groups = defaultdict(list)
    for it in all_items:
        groups[norm(it['description'])].append(it)
    multi = {k: v for k, v in groups.items()
             if len({x['sheet'] for x in v}) >= 3}

    def classify(occs):
        if any(o['cost_total_pretax'] == 0 for o in occs):
            return ('IRREGULAR_OR_TIERED', None)
        if len(occs) < 3:
            return ('SINGLE_EVENT_ONLY', None)
        ratios = []
        for o in occs:
            ev = o.get('event') or {}
            tp = ev.get('total_pax', 0)
            if tp > 0:
                ratios.append(o['cost_total_pretax'] / tp)
        if len(ratios) < 3:
            return ('SINGLE_EVENT_ONLY', None)
        mean_r = sum(ratios) / len(ratios)
        spread = (max(ratios) - min(ratios)) / mean_r if mean_r else 999
        if spread < 0.3:
            return ('PER_PAX_LINEAR', mean_r)
        return ('IRREGULAR_OR_TIERED', None)

    validated = {}
    counts = Counter()
    for desc, occs in multi.items():
        label, ratio_val = classify(occs)
        counts[label] += 1
        if label == 'PER_PAX_LINEAR':
            validated[desc] = ratio_val
    return validated

    validated_t = validated
# ── CELL 13 – templates ───────────────────────────────────────────────────────
def get_tier(total_pax):
    if total_pax < 800:   return 'small'
    if total_pax < 3000:  return 'medium'
    return 'large'

# ── CELL 15 – generator ───────────────────────────────────────────────────────
Fee_CHEFS = {'fee chefs'}
Food_Cost_chefs = {'food cost chefs', 'FOOD COST CHEFS'}
Food_Cost_bites_1 = {'food cost bites 1'}
Food_Cost_bites_2 = {'food cost bites 2'}
Food_Cost_bites_3 = {'food cost bites 3'}
Sexta_Chef = {'chef sexta', 'food cost chefs sexta'}
Sommelier = {'sommelier'}
Fardas = {'fardas'}
Montagens_Desmontagens = {'montagens e desmontagens bares', 'montagens e desmontagens chefs'}
Assistentes_Bares = {'assistentes bares executivo(01 dia antes, evento, 01 dia depois)'}
Sexta = {'sexta almoço', 'sexta jantar'}
Artistas = {'sexta almoço', 'sexta jantar', 'sábado almoço', 'sábado jantar', 'domingo almoço', 'domingo jantar'}
Staff = {'staff'}
Catering_event = {'alimentação staff - durante evento'}
Catering_pre = {'alimentação staff - montagens e desmontagens', 'alimentação staff montagens e desmontagens'}
Catering_band = {'catering backstage bandas', 'bandas'}
Cleaning = {'limpeza'}
Security = {'segurança'}
Equipa = {'equipa montagens média/fina'}
Outros = {'aluguer de viaturas'}
Lenha = {'lenha'}


def generate_budget(event_type, pax_per_day, days, pax_per_slot, slots, month, event_name,
                    extra_notes, templates, validated_ratios, event_meta):
                        
    if event_type == 'fine_dining':
        total_pax_fd = pax_per_slot * slots
        tier_fd = get_tier(total_pax_fd)
        base_type      = (event_type if event_type in templates
                      else min(event_meta.values(),
                               key=lambda m: abs(m['total_pax'] - total_pax_fd))['event_type'])
    elif event_type == 'Pop_Up':
        total_pax_pu = pax_per_slot * slots * days
        tier_pu = get_tier(total_pax_pu)
        base_type      = (event_type if event_type in templates
                      else min(event_meta.values(),
                               key=lambda m: abs(m['total_pax'] - total_pax_pu))['event_type'])
    else:                   
        total_pax      = pax_per_day * days
        tier = get_tier(total_pax)
        base_type      = (event_type if event_type in templates
                      else min(event_meta.values(),
                               key=lambda m: abs(m['total_pax'] - total_pax))['event_type'])
    
    base_items     = templates[base_type]
    base_total_pax = next(m['total_pax'] for m in event_meta.values()
                          if m['event_type'] == base_type)
    generated = []
                        
    for it in base_items:
        key = norm(it['description'])
        row = {
            'category':    it['category'],
            'subcategory': it.get('subcategory'),
            'description': it['description'],
        }

        if it['days'] is None or it['days'] == "" or it['days'] == 0:
            it['days'] = ""  # Keep it empty
        elif it['days'] > 1:
            it['days'] = days

        #copy from template 
        unit_cost_to_use = to_num(it['unit_cost'], 0)
        qty_to_use = to_num(it['qty'], 1)
        days_to_use = to_num(it['days'], 1)
        tag_to_use = 'TEMPLATE_ESTIMATE'
        cost_total_pretax_to_use = to_num(it['cost_total_pretax'], 0)
        #to_num(value, default)

        row.update(unit_cost=unit_cost_to_use, qty=qty_to_use, days=days_to_use,
                   cost_total_pretax=cost_total_pretax_to_use,
                   tag='TEMPLATE_ESTIMATE',
                   note='')
        # Apply specific calculations for 'festival' event_type first
        if event_type == 'festival':
            food_portion_per_day = pax_per_day * 5

            # Convert it['unit_cost'] to a number for internal calculations if it's used
            numeric_it_unit_cost = to_num(it['unit_cost'], 0)
    
            if any(k in key for k in Fee_CHEFS):
                n = math.floor(food_portion_per_day / 239)
                no_of_chef_per_day = math.floor((90 * n) / 1500) + math.floor((60 * n) / 1200) + math.floor((24 * n) / 800) + math.floor((20 * n) / 1000) + math.floor((n * 45) / 1500)
                qty_to_use = no_of_chef_per_day
                days_to_use = days # New event's days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Food_Cost_chefs):
                qty_to_use = food_portion_per_day
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Food_Cost_bites_1) or any(k in key for k in Food_Cost_bites_2) or any(k in key for k in Food_Cost_bites_3):
                food_bites_portion = 2 * pax_per_day
                no_of_restaurant_per_day = math.ceil(food_bites_portion / 600) + 2

                if any(k in key for k in Food_Cost_bites_1):
                    no_chef_a = round(no_of_restaurant_per_day / 18 * 9)
                    qty_to_use = 600 * no_chef_a
                    days_to_use = days
                    tag_to_use = 'DATA_BACKED'
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use 

                elif any(k in key for k in Food_Cost_bites_2):
                    no_chef_b = round(no_of_restaurant_per_day / 18 * 4)
                    qty_to_use = 600 * no_chef_b
                    days_to_use = days
                    tag_to_use = 'DATA_BACKED'
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use 

                elif any(k in key for k in Food_Cost_bites_3):
                    no_chef_c = round(no_of_restaurant_per_day / 18 * 5)
                    qty_to_use = 600 * no_chef_c
                    days_to_use = days
                    tag_to_use = 'DATA_BACKED'
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use # Apply after qty, days are determined

            elif key in Staff:
                cost_staff_1 = 22500 * (1.3 ** (days - 1))
                unit_cost_to_use = cost_staff_1
                tag_to_use = 'DATA_BACKED'

            elif key in Catering_band: 
                numeric_it_unit_cost = unit_cost_to_use 
                unit_cost_to_use = 25
                days_to_use = days 
                qty_to_use = 4
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Catering_event): 
                numeric_it_unit_cost = unit_cost_to_use 
                unit_cost_to_use = 10 
                qty_to_use = 400 
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Cleaning): 
                umeric_it_unit_cost = unit_cost_to_use 
                unit_cost_to_use = 10000
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Security): 
                umeric_it_unit_cost = unit_cost_to_use 
                unit_cost_to_use = 10000
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use
                
                
                    
        if event_type == 'corporate_lunch':
            food_portion = 4 * pax_per_day
            numeric_it_unit_cost = to_num(it['unit_cost'], 0)
            if any(k in key for k in Food_Cost_chefs):
                qty_to_use = food_portion
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Catering_event):  
                if days > 1: 
                    days_to_use = days
                    tag_to_use = 'DATA_BACKED'
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use


        if event_type == 'corporate_dinner':
            food_portion = 9 * pax_per_day
            numeric_it_unit_cost = to_num(it['unit_cost'], 0)
            if any(k in key for k in Fee_CHEFS):
              qty_to_use = 10
              days_to_use = days
              tag_to_use = 'DATA_BACKED'
              cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Food_Cost_chefs):
              qty_to_use = food_portion
              days_to_use = days
              tag_to_use = 'DATA_BACKED'
              cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif key in Staff: 
                unit_cost_to_use = 6300
                if days > 1: 
                    unit_cost_to_use = 6300 * (1.1 ** (days - 1))
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use

            elif key in Catering_band: 
                days_to_use = days 
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * days_to_use

            elif any(k in key for k in Catering_event):  
                if days > 1: 
                    days_to_use = days
                    tag_to_use = 'DATA_BACKED'
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Cleaning):
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Security): 
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

        if event_type == 'fine_dining': # Changed to 'fine_dining' to match event_type in EVENT_META
            numeric_it_unit_cost = to_num(it['unit_cost'], 0)
            no_of_chef = 6
            if any(k in key for k in Fee_CHEFS):
              # Fixed typo: 'desciption' to 'description', 'Fee_Chefs' to 'Fee_CHEFS'
              qty_to_use = no_of_chef
              days_to_use = 1 # Fixed fee, not daily
              # unit_cost_to_use already holds numeric_it_unit_cost
              tag_to_use = 'DATA_BACKED'
              cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Food_Cost_chefs):
                unit_cost_to_use = 80
                numeric_it_unit_cost = unit_cost_to_use
                qty_to_use = total_pax_fd
                days_to_use = 1 # Total pax, not per day
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use
                
            elif any(k in key for k in Sommelier):
                qty_to_use = slots
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use 

            elif any(k in key for k in Fardas):
                qty_to_use = no_of_chef
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use 

            elif key in Staff: 
                qty_to_use = slots
                unit_cost_to_use = 140 * 12 + 70 
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use 
                
            elif any(k in key for k in Catering_event): 
                unit_cost_to_use = 100
                qty_to_use = slots
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use

            elif any(k in key for k in Cleaning): 
                unit_cost_to_use = 13
                qty_to_use = 5
                days_to_use = slots
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use

        if event_type == 'Pop_Up':
            numeric_it_unit_cost = to_num(it['unit_cost'], 0)
            band_fee_est = 3500
            cost_staff = 8000 * (1.1 ** ( days - 1 ))
            if any (k in key for k in Food_Cost_chefs):
                qty_to_use = pax_per_slot * slots * 4
                days_to_use = 1
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use
                if any(k in key for k in Sexta_Chef):
                    qty_to_use = 0 
                    cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use
                    

            elif any(k in key for k in Montagens_Desmontagens):
                days_to_use = 2
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use
                
            elif any(k in key for k in Assistentes_Bares):
                days_to_use = days + 2
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif key in Staff: 
                unit_cost_to_use = cost_staff
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use
    
            elif any(k in key for k in Artistas):
                unit_cost_to_use = band_fee_est
                tag_to_use = 'DATA_BACKED'
                if any(k in key for k in Sexta): 
                    if days <= 2: 
                        unit_cost_to_use = 0 
                        tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use

            elif any(k in key for k in Sexta_Chef):
                if days <= 2: 
                        unit_cost_to_use = 0 
                        tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use 

            elif key in Catering_band: 
                unit_cost_to_use = 25
                days_to_use = slots * days 
                qty_to_use = 2 #band per day 
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use

            elif key in Catering_event: 
                unit_cost_to_use = 16 
                qty_to_use = 120 * slots
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use

            elif key in Catering_pre: 
                days_to_use = days + 4
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Cleaning): 
                unit_cost_to_use = 13
                qty_to_use = 250
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use

            elif key in Security: 
                unit_cost_to_use = 13
                qty_to_use = 205
                days_to_use = days
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * days_to_use

            elif any(k in key for k in Equipa):
                days_to_use = 6
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Outros):
                days_to_use = days + 7 
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use * days_to_use

            elif any(k in key for k in Lenha): 
                qty_to_use = (4 * days) + 4
                tag_to_use = 'DATA_BACKED'
                cost_total_pretax_to_use = numeric_it_unit_cost * qty_to_use

            
        
        row.update(unit_cost=unit_cost_to_use, qty=qty_to_use, days=days_to_use,
                   cost_total_pretax=cost_total_pretax_to_use,
                   tag=tag_to_use,
                   note='')
        
        if tag_to_use == 'TEMPLATE_ESTIMATE':
            matched_ratio = next((rv for rk, rv in validated_ratios.items() if rk in key), None)
            d_for_calc = to_num(it['days'], 1) if it['days'] not in (None, '', 0) else 1
            cost_total_pretax_to_use = unit_cost_to_use * qty_to_use * d_for_calc

            if matched_ratio is not None:
                if event_type == 'fine_dining':
                    unit_cost_to_use = matched_ratio
                    qty_to_use = total_pax_fd
                    days_to_use = 1
                    cost_total_pretax_to_use = matched_ratio * total_pax_fd
                elif event_type == 'Pop_Up':
                    unit_cost_to_use = matched_ratio
                    qty_to_use = total_pax_pu
                    days_to_use = 1
                    cost_total_pretax_to_use = matched_ratio * total_pax_pu
                else: 
                    unit_cost_to_use = matched_ratio
                    qty_to_use = total_pax
                    days_to_use = 1
                    cost_total_pretax_to_use = matched_ratio * total_pax
                tag_to_use = 'DATA_BACKED'
                
            elif to_num(it['cost_total_pretax'], 0) == 0: # Use to_num for comparison
                cost_total_pretax_to_use = 0
                tag_to_use = ('NEEDS_INPUT' if base_type != event_type
                              else 'TEMPLATE_ESTIMATE')

            row.update(unit_cost=unit_cost_to_use, qty=qty_to_use, days=it['days'],
                   cost_total_pretax=cost_total_pretax_to_use,
                   tag=tag_to_use,
                   note='')
            
        generated.append(row)
    
    if event_type == 'fine_dining':
        return {
        'event_name':         event_name,
        'event_type':         event_type,
        'base_template_used': base_type,
        'pax_per_day':        pax_per_day,
        'days':               days,
        'pax_per_slot':       pax_per_slot, 
        'slots':              slots,
        'total_pax':          total_pax_fd,
        'tier':               tier_fd,
        'notes':              extra_notes,
        'line_items':         generated,
        }
    elif event_type == 'Pop_Up':
        return {
        'event_name':         event_name,
        'event_type':         event_type,
        'base_template_used': base_type,
        'pax_per_day':        pax_per_day,
        'days':               days,
        'pax_per_slot':       pax_per_slot, 
        'slots':              slots,
        'total_pax':          total_pax_pu,
        'tier':               tier_pu,
        'notes':              extra_notes,
        'line_items':         generated,
        }
    else:
        return {
        'event_name':         event_name,
        'event_type':         event_type,
        'base_template_used': base_type,
        'pax_per_day':        pax_per_day,
        'days':               days,
        'pax_per_slot':       pax_per_slot, 
        'slots':              slots,
        'total_pax':          total_pax,
        'tier':               tier,
        'notes':              extra_notes,
        'line_items':         generated,
        }
        


# ── CELL 17 – Excel writer ────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', fgColor='434343')
HEADER_FONT   = Font(name='Arial Narrow', size=10, bold=True, color='FFFFFF')
CATEGORY_FILL = PatternFill('solid', fgColor='BAD2D4')
CATEGORY_FONT = Font(name='Arial Narrow', size=10, bold=True, color='000000')
SUBCAT_FONT   = Font(name='Arial Narrow', size=10, bold=True, color='000000')
BASE_FONT     = Font(name='Arial Narrow', size=10)

TAG_FONT = {
    'DATA_BACKED':       Font(name='Arial Narrow', size=10, color='006100'),
    'TEMPLATE_ESTIMATE': Font(name='Arial Narrow', size=10, color='9C6500'),
    'NEEDS_INPUT':       Font(name='Arial Narrow', size=10, color='FF0000', bold=True),
}
TAG_LABEL = {'DATA_BACKED': '', 'TEMPLATE_ESTIMATE': '', 'NEEDS_INPUT': ''}

month_cal = ['JAN', 'FEV', 'MARCO','ABRIL', 'MAIO', 'JUNHO', 'JULHO',
'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

reduced_13 = {'Food Cost chefs', 'Food Cost bites', 'FOOD COST CHEFS'} #EDIT FOR FOOD BITES
reduced_6 = {'Copos'}
extra = {'PALCOS INTERNACIONAIS'}

def write_budget_xlsx(generated) -> bytes:
    """Returns the .xlsx file as bytes so Streamlit can serve it as a download."""
    wb_out = Workbook()
    ws     = wb_out.active
    ws.title = f'wb_{generated['event_type']} '

    ws.merge_cells('A1:F1')
    ws['A1'] = f'BUDGET_{generated['event_type']}'
    ws['A1'].font = Font(name='Arial Narrow', size=12, bold=True)
    ws.merge_cells('G1:M1')
    ws['G1'] = generated['event_name']
    ws['G1'].font = Font(name='Arial Narrow', size=12, bold=True)

    ws.merge_cells('A2:M2')
    event_type = generated['event_type']

    
    if event_type == 'fine_dining':
        meta_txt = (f"Type: {generated['event_type']}  | {generated['slots']} slot(s)  |  "
                f"{generated['pax_per_slot']:.0f} pax/slot |  "
                f"{generated['total_pax']:.0f} total pax  |  "
                f"Scale tier: {generated['tier']}  |  "
                f"Base template: {generated['base_template_used']}")
    elif event_type == 'Pop_Up':
        meta_txt = (f"Type: {generated['event_type']}  |  {generated['days']} day(s)  |  "
                f"{generated['pax_per_day']:.0f} pax/day  | "
                f"{generated['slots']} slot(s)  |  "
                f"{generated['pax_per_slot']:.0f} pax/slot  |  "
                f"{generated['total_pax']:.0f} total pax  |  "
                f"Scale tier: {generated['tier']}  | "
                f"Base template used: {generated['base_template_used']}")
    else:
        meta_txt = (f"Type: {generated['event_type']}  |  {generated['days']} day(s)  |  "
                f"{generated['pax_per_day']:.0f} pax/day  | "
                f"{generated['total_pax']:.0f} total pax  |  "
                f"Scale tier: {generated['tier']}  | "
                f"Base template used: {generated['base_template_used']}")
        
    if generated.get('notes'):
        meta_txt += f"  |  {generated['notes']}"
    ws['A2'] = meta_txt
    ws['A2'].font = Font(name='Arial Narrow', size=9, italic=True)

    event_month = generated.get('month', 'JUNHO').upper() # Default to JUNHO if month is not in generated
    current_month_index = month_cal.index(event_month) if event_month in month_cal else 0

    A = month_cal[(current_month_index - 3 + len(month_cal)) % len(month_cal)]
    B = month_cal[(current_month_index - 2 + len(month_cal)) % len(month_cal)]
    C = month_cal[(current_month_index - 1 + len(month_cal)) % len(month_cal)]
    D = month_cal[(current_month_index + len(month_cal)) % len(month_cal)]


    if event_type == 'Pop_Up':
        COLUMNS = [ ('SITUAÇÃO', 14), ('FORMA DE PAGAMENTO', 16), ('DESCRIÇÃO', 42),
        ('CUSTO UNITÁRIO ESTIMATIVO', 12), ('UNIT', 8), ('DIAS', 7),
        ('CUSTO TOTAL\nESTIMATIVO', 13), ('IVA', 11), ('TOTAL\n(COM IVA)', 13),
        ('PATROCÍNIO', 11), ('ENTIDADE', 15), ('NA ADJUDICAÇÃO(COM IVA)', 15) , ('ATÉ O EVENTO(COM IVA)', 15),
        ('A 30 DIAS',11), ('A 60 DIAS', 11), ('CONFIANÇA / CONFIDENCE', 30),
        ('OBSERVAÇÕES', 40),('valor_total', 40) ]
    else:
        COLUMNS = [
        ('SITUAÇÃO', 14), ('FORMA DE PAGAMENTO', 16), ('DESCRIÇÃO', 42),
        ('CUSTO UNITÁRIO ESTIMATIVO', 12), ('UNIT', 8), ('DIAS', 7),
        ('CUSTO TOTAL\nESTIMATIVO', 13), ('IVA', 11), ('TOTAL\n(COM IVA)', 13),
        ('PATROCÍNIO', 11), ('ENTIDADE', 15), (A + '(COM IVA)', 15), (B + '(COM IVA)', 15), (C + '(COM IVA)', 15),
        (D + '(COM IVA)', 15), ('A 30 DIAS',11), ('A 60 DIAS', 11), ('CONFIANÇA / CONFIDENCE', 30),
        ('OBSERVAÇÕES', 40),('valor_total', 40)]
        
    header_row = 4
    for idx, (label, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(header_row, idx, label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, horizontal='center',
                                vertical='center')
        ws.column_dimensions[get_column_letter(idx)].width = width

    cats = OrderedDict()
    for item in generated['line_items']:
        cats.setdefault(item['category'], []).append(item)

    r = header_row + 1
    grand_total_rows = []

    for cat_name, cat_items in cats.items():
        cat_row = r
        ws.cell(r, 3, cat_name).font = CATEGORY_FONT
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(r, col).fill = CATEGORY_FILL
    
        r += 1
        first_item_row = r
        current_subcat = object()          # sentinel

        for item in cat_items:
            subcat = item.get('subcategory')
            if subcat != current_subcat:
                current_subcat = subcat
                if subcat:
                    ws.cell(r, 3, subcat).font = SUBCAT_FONT
                    r += 1

            ws.cell(r, 3, item['description'])
            ws.cell(r, 4, item['unit_cost'])
            ws.cell(r, 5, item['qty'])
            
            use_days = item['days'] not in (None, 1)
            ws.cell(r, 6, item['days'] if use_days else None)
            ws.cell(r, 7, f"=E{r}*D{r}*F{r}" if use_days else f"=E{r}*D{r}")
            vat = '13%' if item['description'] in reduced_13 else ('6%' if item['description'] in reduced_6 else '23%')
            ws.cell(r, 8, f"=G{r}*{vat}")
            ws.cell(r, 9, f"=G{r}+H{r}")
            
            ws.cell(r, 11, item.get('entity') or '')
            tag_cell      = ws.cell(r, 12, TAG_LABEL[item['tag']])
            tag_cell.font = TAG_FONT[item['tag']]
            ws.cell(r, 13, item.get('note', ''))
            
            for col in (1, 2, 3, 10, 11, 13):
                ws.cell(r, col).font = BASE_FONT
            ws.cell(r, 4).font = TAG_FONT[item['tag']]
            for col in (4, 7, 8, 9):
                ws.cell(r, col).number_format = '#,##0.00'
            r += 1

        last_item_row = r - 1
        if last_item_row >= first_item_row:
            ws.cell(cat_row, 7, f"=SUM(G{first_item_row}:G{last_item_row})")
            ws.cell(cat_row, 8, f"=SUM(H{first_item_row}:H{last_item_row})")
            ws.cell(cat_row, 9, f"=SUM(I{first_item_row}:I{last_item_row})")
            
            for col in (7, 8, 9):
                ws.cell(cat_row, col).font          = CATEGORY_FONT
                ws.cell(cat_row, col).number_format = '#,##0.00'
            grand_total_rows.append((cat_name, cat_row))#changed
        r += 1

    total_row = r + 1
    ws.cell(total_row, 3, 'TOTAL').font = Font(name='Arial Narrow',
                                               size=11, bold=True)
    for col, letter in ((7, 'G'), (8, 'H'), (9, 'I')):
        rng  = '+'.join(f'{letter}{row_num}' for _, row_num in grand_total_rows) #changed
        cell = ws.cell(total_row, col, '=' + rng)
        cell.font           = Font(name='Arial Narrow', size=11, bold=True)
        cell.number_format  = '#,##0.00'
        cell.fill           = PatternFill('solid', fgColor='D9D9D9')

    error_row = total_row + 2
    l_row = error_row + 3
    ws.cell(error_row, 3, 'MARGEM DE ERRO').font = Font(name='Arial Narrow', size=11, bold=True)
    ws.cell(error_row, 5, 'MÁXIMO:').font = Font(name='Arial Narrow', size=11, bold=False)
    ws.cell(error_row + 1, 5, 'MÍNIMO:').font = Font(name='Arial Narrow', size=11, bold=False)

    max = ws.cell(error_row, 7, f'=G{total_row}*1.1')
    max.font = Font(name='Arial Narrow', size=11, bold=True)
    max.number_format = '#,##0.00'
    min = ws.cell(error_row + 1, 7, f'=G{total_row}/1.1')
    min.font = Font(name='Arial Narrow', size=11, bold=True)
    min.number_format = '#,##0.00'

    #add income part
    inc_header_row = l_row + 1
    ws.cell(inc_header_row, 3, 'INCOME').font = Font(name='Arial Narrow', size=11, bold=True)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(inc_header_row, col).fill = CATEGORY_FILL

    in_items = ['Patrocinio 1', 'Patrocinio 2', 'Patrocinio 3', 'Patrocinio 4', 'Patrocinio 5',
                'Patrocinio 6', 'Bilheteira', 'Receita Bares']

    current_income_item_row = inc_header_row + 1
    for item_text in in_items:
        ws.cell(current_income_item_row, 3, item_text).font = Font(name='Arial Narrow', size=11, bold=False)
        current_income_item_row += 1

    inc_ttl_r = current_income_item_row + 1
    inc_final_r = inc_ttl_r + 2
    ws.cell(inc_ttl_r, 3, 'TOTAL').font = Font(name='Arial Narrow', size=11, bold=True, color='FFFFFF')
    ws.cell(inc_final_r, 3, 'SALDO FINAL').font = Font(name='Arial Narrow', size=11, bold=True, color='FFFFFF')
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(inc_ttl_r, col).fill = PatternFill('solid', fgColor='595959')
        ws.cell(inc_final_r, col).fill = PatternFill('solid', fgColor='8A9F8E')

    last_inc_row = inc_final_r + 1 

    options_situacao = [
          'VALOR ESTIMADO',
          'ORÇAMENTADO',
          'ADJUDICADO (A AGUARDAR FATURA)',
          'FALTA MAIS PAGAMENTOS',
          'PAGAMENTO EM FALTA',
          'PAGO',
          'N/A']
    options_forma_de_pagamento = [
          'TRANSFERÊNCIA',
          'REVOLUT JHONATHA',
          'REVOLUT LAUREANO',
          'REVOLUT VASCO',
          'SPONSOR',
          'NUMERÁRIO',
          'N/A',
          'REEMBOLSO']
    options_patrocinio = [
          'Sim',
          'Não' ]

    #Data Validation required
    situacao_dv = DataValidation(
          type = "list",
          formula1=f'"{",".join(options_situacao)}"',
          allow_blank=True)

    forma_de_pagameto_dv = DataValidation(
          type = "list",
          formula1=f'"{",".join(options_forma_de_pagamento)}"',
          allow_blank=True)


    patrocinio_dv = DataValidation(
            type = "list",
            formula1=f'"{", ".join(options_patrocinio)}"',
            allow_blank=True)


    ws.add_data_validation(situacao_dv)
    ws.add_data_validation(forma_de_pagameto_dv)
    ws.add_data_validation(patrocinio_dv)

    # Determine the full range for data validation
    start_row = header_row + 2
    end_row = r - 1 # 'r' is one past the last item row, so r-1 is the last item row.

    if start_row <= end_row: # Only apply if there are actual item rows
        situacao_dv.add(f'{get_column_letter(1)}{start_row}:{get_column_letter(1)}{end_row}')
        forma_de_pagameto_dv.add(f'{get_column_letter(2)}{start_row}:{get_column_letter(2)}{end_row}')
        patrocinio_dv.add(f'{get_column_letter(10)}{start_row}:{get_column_letter(10)}{end_row}')

    
    

    # Add a pie chart for category breakdown
    chart_start_row = last_inc_row + 3

    # 1. Create a summary table for the chart data
    summary_header_row = chart_start_row
    ws.cell(summary_header_row, 2, "Ranking").font = Font(bold=True)
    ws.cell(summary_header_row, 3, "Category").font = Font(bold=True)
    ws.cell(summary_header_row, 4, "Total (Com IVA)").font = Font(bold=True)
    ws.cell(summary_header_row, 5, "Weighting").font = Font(bold=True)

    # --- Python-side calculation for accurate ranking ---
    category_actual_totals = defaultdict(float)
    grand_total_actual = 0.0

    for item in generated['line_items']:
        # Determine VAT rate
        vat_rate_factor = 0.23 # Default
        if item['description'] in reduced_13: vat_rate_factor = 0.13
        elif item['description'] in reduced_6: vat_rate_factor = 0.06

        item_total_pretax = item['cost_total_pretax']
        item_total_com_iva = item_total_pretax * (1 + vat_rate_factor)

        # Apply 'extra' multipliers if applicable (from the budget logic)
        # if item['category'] in extra:
        #     item_total_pretax *= 3 # This is 'v' from earlier
        #     item_total_com_iva *= 5 # This is 'c' from earlier

        category_actual_totals[item['category']] += item_total_com_iva

    # Calculate grand total from Python-calculated category totals
    grand_total_actual = sum(category_actual_totals.values())

    # Prepare data for sorting
    rankable_categories = []
    for cat_name, total_com_iva_actual in category_actual_totals.items():
        weighting = (total_com_iva_actual / grand_total_actual) * 100 if grand_total_actual > 0 else 0

        # Find the original Excel row number for this category from grand_total_rows
        excel_total_row_for_category = None
        for name, r_num in grand_total_rows:
            if name == cat_name:
                excel_total_row_for_category = r_num
                break

        rankable_categories.append({
            'Category': cat_name,
            'Total (Com IVA)': total_com_iva_actual, # Actual Python calculated value (for sorting)
            'Weighting (%)': weighting,
            'Excel_Total_Row': excel_total_row_for_category # Original Excel row for formula reference
        })

    # Sort categories by Weighting (%) in descending order
    rankable_categories_sorted = sorted(rankable_categories, key=lambda x: x['Weighting (%)'], reverse=True)

    current_summary_row = summary_header_row + 1
    for rank_idx, ranked_cat_info in enumerate(rankable_categories_sorted, start=1):
        ws.cell(current_summary_row, 2, rank_idx).font = BASE_FONT # Rank column
        ws.cell(current_summary_row, 3, ranked_cat_info['Category']).font = BASE_FONT # Category name

        # Write Excel formulas that reference the original category total cells
        if ranked_cat_info['Excel_Total_Row'] is not None:
            ws.cell(current_summary_row, 4, f'=I{ranked_cat_info['Excel_Total_Row']}').number_format = '#,##0.00'
            ws.cell(current_summary_row, 5, f'=I{ranked_cat_info['Excel_Total_Row']}/I{total_row}*100').number_format = '#,##0.00'
        else:
             ws.cell(current_summary_row, 4, 0).number_format = '#,##0.00'
             ws.cell(current_summary_row, 5, 0).number_format = '#,##0.00'

        current_summary_row += 1

    summary_data_end_row = current_summary_row - 1 # Update end row after writing sorted data

    # 2. Create the PieChart using this new summary table
    chart = PieChart()
    chart.title = "Cost breakdown of main categories"
    chart.height = 32
    chart.width = 40

    # Categories (labels) are now in column 3 of the summary table
    labels_ref = Reference(ws, min_col=3, min_row=summary_header_row + 1, max_row=summary_data_end_row)
    # Data (values) are now in column 5 of the summary table (Weighting)
    data_ref = Reference(ws, min_col=5, min_row=summary_header_row + 1, max_row=summary_data_end_row)

    series = Series(data_ref)
    chart.series.append(series)
    chart.set_categories(labels_ref)

    # Add data labels with percentage
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.numFmt = '#.##%'
    chart.dataLabels.showCatName = True
    chart.dataLabels.showSeriesName = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showVal = False
    chart.dataLabels.dLblPos = "outEnd"

    # Position the chart right after the summary table (e.g., starting from column F)
    ws.add_chart(chart, f"F{summary_header_row}")

    ws.freeze_panes = 'A5'
    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()

# ── CELL 19 – NL extractor ────────────────────────────────────────────────────
EVENT_TYPE_KEYWORDS = {
    'Pop-Up' : ['Pop-up', 'Pop up', 'pop up', 'pop-up', 'POPUP', 'pop up event', 'Pop Up Event'],
    'Chefs on Fire': ['festival', 'Cascais', 'food festival', 'Chefs On Fire', 'COF'],
    'Almoço Corp.': ['corporate lunch', 'Almoço Corp', 'Almoço', 'Almoco Corp', 'Almoco','lunch event'],
    'Jantar Corp.': ['corporate dinner','dinner event','Jantar Corp', 'Jantar' ],
    'Fine Dining': ['fine dining', 'Fine Dining', 'dining']
}

def extract_params(text):
    tl = text.lower()
    event_type = next((et for et, kws in EVENT_TYPE_KEYWORDS.items()
                       if any(k in tl for k in kws)), 'festival')
    dm = re.search(r'(\d+)\s*-?\s*day', tl)
    days = int(dm.group(1)) if dm else 1
    pm = re.search(r'(\d[\d,.]*)\s*(?:pax|people|participants|guests|attendees)', tl)
    pax = int(pm.group(1).replace(',','').replace('.','')) if pm else None
    sm = re.search(r'(\d+)\s*-?\s*slot', tl)
    slots = int(sm.group(1)) if sm else 1
    pm_s = re.search(r'(\d[\d,.]*)\s*(?:pax|people|participants|guests|attendees)\s*per\s*slot', tl)
    pax_s = int(pm_s.group(1).replace(',','').replace('.','')) if pm_s else None
    date_match = re.search(r'on\s+[[[\[\d\s\-.]*(\w+)', text_lower)
    month = date_match.group(1).upper() if date_match else 'JUNHO'
    nm = re.search(r'(?:called|named|titled)\s*["\']?([\w\s\-]+)["\']?',
                   text, re.IGNORECASE)
    name = nm.group(1).strip() if nm else 'New Event'
    return {'event_type': event_type, 'days': days,
            'pax_per_day': pax, 'slots': slots, 'pax_per_slot': pax_s, 'month': month, 'event_name': name}

# ═════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═════════════════════════════════════════════════════════════════════════════

# ── EVENT_META: edit sheet names here to match your Excel tabs ────────────────
# Defined at module level (not inside the sidebar block) so it is always
# available before parsing runs. This was the root cause of the original error.
EVENT_META = {
    'wb_Cascais': {
        'event_name': "Chef's on Fire – Cascais 2026",
        'event_type': 'festival',
        'days': 2, 'pax_per_day': 5000,
    },
    'wb_Almoço Corp.': {
        'event_name': "Chef's on Fire – Almoço Corporate",
        'event_type': 'corporate_lunch',
        'days': 1, 'pax_per_day': 500,
    },
    'wb_Jantar Corp.': {
        'event_name': "Chef's on Fire – Dia Corporate",
        'event_type': 'corporate_dinner',
        'days': 1, 'pax_per_day': 1000,
    },
    'wb_Fine Dining': {
        'event_name': "Chef's on Fire – Fine Dining",
        'event_type': 'fine_dining',
        'days': 8, 'slots' : 6, 'pax_per_slot': 24,
    },
    'wb_Pop-up': {
        'event_name': "Chef's on Fire – Pop-up",
        'event_type': 'Pop_Up',
        'days': 3, 'slots': 2, 'pax_per_slot': 600,
    },
}


for m in EVENT_META.values(): 
    if m['event_type'] == 'fine_dining':
        m['total_pax'] = m['pax_per_slot'] * m['slots']
    elif m['event_type'] == 'Pop_Up':
        m['total_pax'] = m['pax_per_slot'] * m['slots'] * m['days']
    else: 
        m['total_pax'] = m['pax_per_day'] * m['days']

# ── sidebar: show template info ───────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Event templates")
    st.caption("These must match the exact sheet names in your uploaded Excel. "
               "Edit them here if your file uses different names.")

    for sheet, meta in EVENT_META.items():
        st.markdown(f"**`{sheet}`** → `{meta['event_type']}`  ")

    st.markdown("---")
    st.caption("Confidence colour key in generated Excel:\n"
               "🟢 Green = formula-backed\n"
               "🟠 Amber = copied estimate — verify\n"
               "🔴 Red = needs your input")

# ── Step 1: upload ────────────────────────────────────────────────────────────
st.markdown("### 1 · Upload your historical budget Excel")
st.markdown(
    '<div class="notice">Upload the same workbook you\'ve always used '
    '(e.g. <code>Budget_Cascais_26.xlsx</code>). It is read into memory '
    'and never stored anywhere.</div>',
    unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Choose your .xlsx file", type="xlsx", label_visibility="collapsed")

if not uploaded_file:
    st.info("Upload your budget Excel above to get started.")
    st.stop()

# ── Parse the workbook (same as CELL 9 loop) ─────────────────────────────────
@st.cache_data(show_spinner="Reading historical budgets…")
def load_data(file_bytes: bytes, event_meta: str):
    """Cache is keyed on raw bytes + meta string so a new file invalidates it."""
    meta = json_loads(event_meta)
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_items = []
    warnings  = []
    for sheet_name, m in meta.items():
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet **{sheet_name}** not found — skipped.")
            continue
        items = parse_sheet(wb[sheet_name], sheet_name)
        for it in items:
            it['event'] = m
        all_items.extend(items)
    return all_items, warnings

import json as _json
def json_loads(s): return _json.loads(s)

file_bytes  = uploaded_file.read()
meta_key    = _json.dumps(
    {k: {kk: vv for kk, vv in v.items() if kk != 'total_pax'}
     for k, v in EVENT_META.items()}, sort_keys=True)

all_items, load_warnings = load_data(file_bytes, meta_key)
for w in load_warnings:
    st.warning(w)

if not all_items:
    st.error("No line items could be parsed. Check that the sheet names in "
             "the sidebar exactly match those in your uploaded file.")
    st.stop()

# Build templates and validated ratios (CELL 11 + 13)
templates        = defaultdict(list)
for it in all_items:
    templates[it['event']['event_type']].append(it)
validated_ratios = build_validated_ratios(all_items)

st.success(
    f"Loaded **{len(all_items)}** historical line items across "
    f"**{len(templates)}** event type(s). "
    f"Validated per-pax ratios: **{len(validated_ratios)}** "
    f"({'none yet — normal with < 30 events' if not validated_ratios else ', '.join(validated_ratios)})."
)

# ── Step 2: describe the new event ───────────────────────────────────────────
st.markdown("### 2 · Describe the new event")

col_l, col_r = st.columns([2, 3])
with col_l:
    user_text = st.text_area(
        "Plain-language description",
        placeholder='Please enter extra info. (e.g. 11 setembro - 18h00 às 23h00)',
        height=90,
        label_visibility="collapsed",
     )
with col_r:
    st.caption("The details of your event")
    manual_type = st.selectbox(
        "Event type", options=list(EVENT_TYPE_KEYWORDS.keys()), index=0)
    manual_days = st.number_input("Days", min_value=1, value=2)
    manual_pax  = st.number_input("Pax per day (🔴Skip if not applicable)", min_value=1, value=2000)
    manual_slots = st.number_input("Slot per day/Total Number of Slots (🔴 Skip if not applicable)", min_value=1, value=2)
    manual_pax_slot = st.number_input("Pax per slot(🔴 Skip if not applicable)", min_value=1, value=700)
    manual_month = st.selectbox ("Date of the event", options=month_cal, index=0)
    manual_name = st.text_input("Event name", value="New Event 2027")

use_manual = not user_text.strip()

    
    
# ── Step 3: generate ─────────────────────────────────────────────────────────
st.markdown("### 3 · Generate")

if st.button("🔥 Generate budget"):
    if use_manual:
        if manual_type == 'Pop-Up': 
            event_type = 'Pop_Up'
        elif manual_type == 'Chefs on Fire':
            event_type = 'festival'
        elif manual_type == 'Almoço Corp.':
            event_type = 'corporate_lunch'
        elif manual_type == 'Jantar Corp.':
            event_type = 'corporate_dinner'
        elif manual_type == 'Fine Dining':
            event_type = 'fine_dining'
        else: 
            event_type = manual_type
            
       params = {
            'event_type': event_type,
            'days':        manual_days,
            'pax_per_day': manual_pax,
            'slots': manual_slots, 
            'pax_per_slot': manual_pax_slot, 
            'month': manual_month,
            'event_name':  manual_name, }
        


    if params['event_type'] not in templates:
        st.error(
                f"No historical template found for event type "
                f"**{params['event_type']}**. "
                f"Available types: {', '.join(templates.keys())}. "
                "Check the sheet names in the sidebar."
            )
        st.stop()
    
    with st.spinner("Building budget…"):
        result     = generate_budget(
                event_type       = params['event_type'],
                pax_per_day      = params['pax_per_day'],
                days             = params['days'],
                pax_per_slot = params['pax_per_slot'],
                slots = params['slots'],
                month = params['month'], 
                event_name       = params['event_name'],
                extra_notes      = user_text,
                templates        = templates,
                validated_ratios = validated_ratios,
                event_meta       = EVENT_META,
            )
        xlsx_bytes = write_budget_xlsx(result)

    # ── Results ───────────────────────────────────────────────────────────────
    tags = Counter(it['tag'] for it in result['line_items'])
    n    = len(result['line_items'])

    st.success(f"Done — **{n} line items** across "
               f"**{len({it['category'] for it in result['line_items']})}** categories.")

    m1, m2, m3 = st.columns(3)
    m1.metric("🟢 Formula-backed",  tags.get('DATA_BACKED', 0))
    m2.metric("🟠 Estimate",         tags.get('TEMPLATE_ESTIMATE', 0))
    m3.metric("🔴 Needs your input", tags.get('NEEDS_INPUT', 0))

    # Category cost preview
    st.markdown("**Category totals (incl. IVA) — matches Excel column I**")
    cat_totals = defaultdict(float)
    for it in result['line_items']:
        # Mirror the Excel formula: col G (pre-tax) × (1 + VAT rate) = col I
        #vat = 0.13 if it['description'] in reduced_13 else (0.06 if it['description'] in reduced_6 else 0.23)
        cat_totals[it['category']] += it['cost_total_pretax'] 
    table_rows = [
        {"Category": k, "Total incl. IVA (€)": f"{v:,.0f}"}
        for k, v in sorted(cat_totals.items(), key=lambda x: -x[1])
    ]
    st.table(table_rows)

    # Download button (replaces files.download from CELL 23)
    fname = f"Budget_{params['event_name'].replace(' ', '_')}.xlsx"
    st.download_button(
        label     = "⬇ Download Excel",
        data      = xlsx_bytes,
        file_name = fname,
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
